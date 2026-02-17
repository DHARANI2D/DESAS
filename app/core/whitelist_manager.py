import json
import os
from app.core.config import settings
from app.core.settings_manager import DATA_DIR

WHITELIST_FILE = os.path.join(DATA_DIR, "whitelist.json")

def get_whitelist() -> list[str]:
    """
    Returns the current whitelist.
    Merges configuration defaults with the persistent JSON file.
    """
    # Start with defaults from config
    current_list = set(settings.DOMAIN_WHITELIST)
    
    if os.path.exists(WHITELIST_FILE):
        try:
            with open(WHITELIST_FILE, 'r') as f:
                saved_list = json.load(f)
                if isinstance(saved_list, list):
                    current_list.update(saved_list)
        except Exception:
            pass # Ignore file errors, fallback to defaults
            
    return sorted(list(current_list))

def add_to_whitelist(domain: str) -> list[str]:
    """
    Adds a domain to the persistent whitelist.
    Returns the updated list.
    """
    current = get_whitelist()
    domain = domain.lower().strip()
    
    if domain and domain not in current:
        current.append(domain)
        current.sort()
        
        # Save ONLY the diff or the whole thing? 
        # For simplicity, we save the entire merged list to the JSON file
        # This means the JSON file becomes the source of truth + config defaults are just seeds
        try:
            with open(WHITELIST_FILE, 'w') as f:
                json.dump(current, f, indent=2)
        except Exception as e:
            print(f"Error saving whitelist: {e}")
            
    return current

def remove_from_whitelist(domain: str) -> list[str]:
    """
    Removes a domain from the persistent whitelist.
    Returns the updated list.
    """
    current = get_whitelist()
    domain = domain.lower().strip()
    
    if domain in current:
        current.remove(domain)
        current.sort()
        
        try:
            with open(WHITELIST_FILE, 'w') as f:
                json.dump(current, f, indent=2)
        except Exception as e:
            print(f"Error saving whitelist: {e}")
            
    return current

def export_whitelist_to_excel(file_path: str):
    """Exports the current whitelist to an Excel file."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Whitelist"
    ws.append(["Domain"])
    
    current = get_whitelist()
    for domain in current:
        ws.append([domain])
    
    wb.save(file_path)
    return True

def import_whitelist_from_excel(file_path: str) -> list[str]:
    """Imports domains from an Excel file into the whitelist. Skips duplicates."""
    import openpyxl
    if not os.path.exists(file_path):
        return get_whitelist()
        
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    new_domains = []
    # Assume first column has domains, skip header if it says 'Domain'
    first_row = True
    for row in ws.iter_rows(values_only=True):
        if first_row and row[0] and str(row[0]).lower() == "domain":
            first_row = False
            continue
        if row[0]:
            new_domains.append(str(row[0]).lower().strip())
    
    current = set(get_whitelist())
    current.update(new_domains)
    
    updated_list = sorted(list(current))
    try:
        with open(WHITELIST_FILE, 'w') as f:
            json.dump(updated_list, f, indent=2)
    except Exception as e:
        print(f"Error saving imported whitelist: {e}")
        
    return updated_list
